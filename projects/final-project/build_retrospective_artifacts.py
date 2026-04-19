from __future__ import annotations

from pathlib import Path
from typing import Iterable

import pdfplumber
from openpyxl import load_workbook
from PIL import Image, ImageDraw, ImageFont


WORKSPACE_DIR = Path(__file__).resolve().parents[1]
OUTPUT_DIR = Path(__file__).resolve().parent / "generated"
SOURCE_DIR = OUTPUT_DIR / "source_extracts"
MARKDOWN_OUTPUT = OUTPUT_DIR / "retrospective_board_package_ua.md"
IMAGE_OUTPUT = OUTPUT_DIR / "retrospective_board_mockup_ua.jpg"


BOARD_METADATA = {
    "title": "Фінальний retrospective board",
    "subtitle": (
        "Ретроспектива 2024-2026: не зміна кар'єри з нуля, а звуження і посилення "
        "траєкторії від broad full-stack/data напряму до AI Integrations / LLMOps / Data Infrastructure."
    ),
    "current_base": "Full-Stack Developer у Full-Iron Design з переходом у MLOps",
    "old_point_a": "Junior MLOps Engineer / Full-Stack Developer (Full-Iron Design)",
    "old_point_b": "MLOps Engineer / Junior AI Engineer",
    "older_pdf_goal": (
        "Full-Stack Data Engineer з довгостроковою амбіцією вирости до менеджерської / Product Owner ролі"
    ),
    "updated_point_b": (
        "AI Integrations / LLMOps / Data Infrastructure Engineer з фокусом на orchestration, "
        "automation, ML/LLM workflows і production-ready delivery"
    ),
    "narrative": (
        "PDF показує ранню широку ціль навколо Full-Stack Data Engineer, автономії, стабільності і growth у Product Owner. "
        "XLSX уже звужує фокус до MLOps Engineer / Junior AI Engineer та фіксує skill gaps: cloud for ML, "
        "Kubernetes, CI/CD for ML, LLMOps, Terraform. Поточні 2026 матеріали показують реальний прогрес: automation з "
        "20% економією ручної праці, scraper/ETL з ефектом понад 7 млн грн, MSc in Data Science, thesis track і чіткий інтерес "
        "до orchestration, AI integrations, data infrastructure та applied AI use cases."
    ),
}


INTERMEDIATE_GOALS = [
    "До 30.06.2026: один production-grade portfolio case у напрямі AI integrations / data workflow з репозиторієм, схемою і short demo.",
    "До 31.07.2026: один cloud або Kubernetes deployment для ML/LLM чи data workflow з IaC і відтворюваним запуском.",
    "До 31.05.2026: один публічний proof point з thesis або agentic workflow експерименту у форматі post / case note / demo.",
    "До 15.05.2026: синхронізований package з LinkedIn, self-presentation і двома CV-версіями під AI Integrations / LLMOps / Data Infrastructure.",
    "На 3 місяці підряд: стабільний pipeline 15-20 цільових відгуків щомісяця з трекінгом фідбеку та корекцією позиціонування.",
]


PRIMARY_ATTACHMENTS = [
    {
        "path": str(WORKSPACE_DIR / "Формулювання кар'єрних цілей.pdf"),
        "purpose": "Раннє формулювання Point A / Point B і ціннісної рамки: автономія, стабільність, служіння.",
    },
    {
        "path": str(WORKSPACE_DIR / "workshop - How to get a raise" / "Illya_Fefelov_SWOT_IPR_filled.xlsx"),
        "purpose": "Точне джерело старої точки A/B, skill gaps і початкового ІПР; саме на нього спиратись для старих формулювань.",
    },
]


SUPPORTING_REFERENCES = [
    str(WORKSPACE_DIR / "professional_interests_understanding_2026.md"),
    str(WORKSPACE_DIR / "resume_illya_fefelov_mlops_2026_revised.md"),
    str(WORKSPACE_DIR / "linkedin_illya_fefelov_2026.md"),
    str(WORKSPACE_DIR / "self_presentation_script.md"),
]


SECTIONS = [
    {
        "title": "Appreciation",
        "fill": "#FFE599",
        "header": "#E0B94B",
        "notes": [
            "10+ років у delivery та full-stack дали не лише технічну базу, а production-thinking: я вмію доводити системи до реального використання, а не лише до демо.",
            "У Full-Iron я вже будував automation і data workflows з вимірюваним ефектом: зниження ручної праці на 20%.",
            "Scraper/ETL ініціатива з ефектом понад 7 млн грн нетто показує, що моя сила не в абстрактному інтересі до AI, а в інженерії з бізнес-результатом.",
            "Перехід не декларативний: MSc in Data Science, thesis track, Python/ML стек і постійна робота з workflows уже змінили профіль у бік MLOps / AI engineering.",
            "Сильна сторона профілю - поєднання full-stack, automation, data pipelines і AI tooling. Саме це робить рух у AI Integrations / LLMOps логічним і доказовим.",
        ],
    },
    {
        "title": "Continue Doing",
        "fill": "#CDECCF",
        "header": "#86C68C",
        "notes": [
            "Продовжувати будувати narrative навколо orchestration, automation, AI integrations і data infrastructure, а не навколо загального full-stack бренду.",
            "Продовжувати використовувати full-stack бекграунд як перевагу для end-to-end AI systems: UI/API, інтеграції, delivery і deployment.",
            "Продовжувати перетворювати навчання в докази: thesis, pet-проєкти, data/ML workflows, agentic tooling, а не лише завершені курси.",
            "Продовжувати орієнтуватися на домени, де є реальні операційні дані та вимірюваний ефект: enterprise, energy, decision-support, AI-enabled products.",
        ],
    },
    {
        "title": "Stop doing",
        "fill": "#F7C7C7",
        "header": "#E38B8B",
        "notes": [
            "Не позиціонувати себе надто широко: full-stack + data + AI + design без центральної осі виглядає як розпорошення, а не як стратегія.",
            "Не залишати публічний образ на рівні люблю AI; без workflows, infra, automation і deployment це звучить занадто загально.",
            "Не відкладати публічну упаковку сильних кейсів до ідеального моменту; зараз саме видимість доказів є вузьким місцем.",
            "Не робити сертифікацію самоціллю: без deployed case вона не замінить production proof.",
        ],
    },
    {
        "title": "Start doing",
        "fill": "#CFE2F3",
        "header": "#7FAFDD",
        "notes": [
            "Почати системно оформлювати 1-2 публічні кейси в стилі AI Integrations / Data Infrastructure: проблема, архітектура, стек, результат.",
            "Почати показувати thesis і agentic workflow експерименти як доказ orchestration-thinking, а не залишати їх лише в навчальному контексті.",
            "Почати будувати cloud / Kubernetes layer навколо одного реального AI або data workflow, щоб закрити старий skill gap з XLSX через артефакт, а не теорію.",
            "Почати регулярний application pipeline під цільові ролі: AI Integrations, LLMOps, Data Infrastructure, MLOps-adjacent product roles.",
            "Почати короткі публічні технічні розбори українською або англійською про automation, ETL, agent workflows і practical AI use cases.",
        ],
    },
    {
        "title": "Actions",
        "fill": "#F9D5A7",
        "header": "#E7A85A",
        "notes": [
            "Portfolio case -> зібрати один публічний AI integration / data workflow кейс з README, схемою, репо і short demo -> готовий артефакт до 30.06.2026.",
            "Cloud / infra closure -> задеплоїти один ML/LLM або data workflow у cloud чи Kubernetes з IaC і reproducible run -> готовий deploy до 31.07.2026.",
            "Public positioning -> синхронізувати CV, LinkedIn, self-presentation і headline під AI Integrations / LLMOps / Data Infrastructure -> один узгоджений narrative до 15.05.2026.",
            "Thesis as proof -> оформити thesis або agentic workflow сегмент як public case note, post або mini-demo -> один proof point до 31.05.2026.",
            "Market loop -> вести трекер 15-20 цільових відгуків щомісяця та фіксувати фідбек -> три місяці підряд без провалу темпу.",
        ],
    },
]


SUBMISSION_CHECKLIST = [
    "Заповнити всі 5 секцій на FigJam: Appreciation, Continue Doing, Stop doing, Start doing, Actions.",
    "Додати на дошку 2 file cards / links: PDF з кар'єрними цілями і XLSX зі SWOT & ІПР.",
    "Вставити формулювання оновленого Point B без скорочень і без абстрактного формату люблю AI.",
    "Додати 5 проміжних цілей з дедлайнами або вимірюваним результатом.",
    "Перевірити, що всі стікери читаються на одному екрані або в одному експорті.",
    "Якщо FigJam буде доступний, експортувати фінальну дошку в JPG і перевірити, що видно title, updated Point B, attachment cards і всі секції.",
]


def ensure_output_dirs() -> None:
    SOURCE_DIR.mkdir(parents=True, exist_ok=True)


def extract_pdf_text(pdf_path: Path, max_pages: int = 6) -> str:
    lines: list[str] = []

    with pdfplumber.open(pdf_path) as pdf:
        lines.append(f"FILE: {pdf_path.name}")
        lines.append(f"PAGES: {len(pdf.pages)}")
        lines.append("")

        for page in pdf.pages[:max_pages]:
            lines.append(f"=== PAGE {page.page_number} ===")
            lines.append(page.extract_text(layout=True) or "")
            lines.append("")

    return "\n".join(lines).strip() + "\n"


def extract_workbook_text(workbook_path: Path) -> str:
    lines: list[str] = [f"FILE: {workbook_path.name}", ""]
    workbook = load_workbook(filename=workbook_path, read_only=True, data_only=True)

    try:
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            lines.append(f"=== SHEET: {sheet_name} ===")

            for row in sheet.iter_rows(values_only=True):
                values = [str(value).strip() for value in row if value not in (None, "")]
                if values:
                    lines.append(" | ".join(values))

            lines.append("")
    finally:
        workbook.close()

    return "\n".join(lines).strip() + "\n"


def load_font(size: int, bold: bool = False) -> ImageFont.FreeTypeFont | ImageFont.ImageFont:
    font_paths = [
        Path("C:/Windows/Fonts/segoeuib.ttf" if bold else "C:/Windows/Fonts/segoeui.ttf"),
        Path("C:/Windows/Fonts/arialbd.ttf" if bold else "C:/Windows/Fonts/arial.ttf"),
    ]

    for font_path in font_paths:
        if font_path.exists():
            return ImageFont.truetype(str(font_path), size=size)

    return ImageFont.load_default()


def wrap_text(draw: ImageDraw.ImageDraw, text: str, font: ImageFont.FreeTypeFont | ImageFont.ImageFont, max_width: int) -> str:
    wrapped_lines: list[str] = []

    for paragraph in text.split("\n"):
        if not paragraph.strip():
            wrapped_lines.append("")
            continue

        words = paragraph.split()
        current_line: list[str] = []

        for word in words:
            test_line = " ".join(current_line + [word])
            if not current_line or draw.textlength(test_line, font=font) <= max_width:
                current_line.append(word)
            else:
                wrapped_lines.append(" ".join(current_line))
                current_line = [word]

        if current_line:
            wrapped_lines.append(" ".join(current_line))

    return "\n".join(wrapped_lines)


def draw_text_block(
    draw: ImageDraw.ImageDraw,
    text: str,
    font: ImageFont.FreeTypeFont | ImageFont.ImageFont,
    fill: str,
    box: tuple[int, int, int, int],
    spacing: int = 8,
) -> None:
    x0, y0, x1, _ = box
    wrapped = wrap_text(draw, text, font, x1 - x0)
    draw.multiline_text((x0, y0), wrapped, font=font, fill=fill, spacing=spacing)


def draw_card(
    draw: ImageDraw.ImageDraw,
    box: tuple[int, int, int, int],
    title: str,
    body_lines: Iterable[str],
    fill: str,
    outline: str,
    title_font: ImageFont.FreeTypeFont | ImageFont.ImageFont,
    body_font: ImageFont.FreeTypeFont | ImageFont.ImageFont,
) -> None:
    x0, y0, x1, y1 = box
    shadow_box = (x0 + 10, y0 + 10, x1 + 10, y1 + 10)
    draw.rounded_rectangle(shadow_box, radius=28, fill="#D9D4C8")
    draw.rounded_rectangle(box, radius=28, fill=fill, outline=outline, width=3)

    title_y = y0 + 24
    draw.text((x0 + 28, title_y), title, font=title_font, fill="#23313D")

    current_y = title_y + 58
    for body_line in body_lines:
        wrapped = wrap_text(draw, body_line, body_font, x1 - x0 - 56)
        draw.multiline_text((x0 + 28, current_y), wrapped, font=body_font, fill="#2E3C46", spacing=7)
        text_box = draw.multiline_textbbox((x0 + 28, current_y), wrapped, font=body_font, spacing=7)
        current_y = text_box[3] + 16


def draw_sticky_section(
    draw: ImageDraw.ImageDraw,
    box: tuple[int, int, int, int],
    title: str,
    notes: list[str],
    header_fill: str,
    note_fill: str,
    title_font: ImageFont.FreeTypeFont | ImageFont.ImageFont,
    body_font: ImageFont.FreeTypeFont | ImageFont.ImageFont,
) -> None:
    x0, y0, x1, y1 = box
    header_height = 74
    gap = 18

    draw.rounded_rectangle((x0, y0, x1, y1), radius=30, fill="#F6F1E8", outline="#D4CBBE", width=2)
    draw.rounded_rectangle((x0 + 4, y0 + 4, x1 - 4, y0 + header_height), radius=26, fill=header_fill)
    draw.text((x0 + 24, y0 + 18), title, font=title_font, fill="#1E2D36")

    available_height = y1 - y0 - header_height - 36 - gap * (len(notes) - 1)
    note_height = int(available_height / max(len(notes), 1))
    note_y = y0 + header_height + 18

    for note in notes:
        note_box = (x0 + 18, note_y, x1 - 18, note_y + note_height)
        shadow_box = (note_box[0] + 8, note_box[1] + 8, note_box[2] + 8, note_box[3] + 8)
        draw.rounded_rectangle(shadow_box, radius=22, fill="#DDD5C8")
        draw.rounded_rectangle(note_box, radius=22, fill=note_fill, outline="#C9B58C", width=2)
        wrapped = wrap_text(draw, note, body_font, note_box[2] - note_box[0] - 32)
        text_y = note_box[1] + 18
        draw.multiline_text((note_box[0] + 16, text_y), wrapped, font=body_font, fill="#2A3740", spacing=6)
        note_y += note_height + gap


def write_board_markdown(output_path: Path) -> None:
    lines = [
        "# Пакет для фінального FigJam retrospective board",
        "",
        "## Рамка ретроспективи",
        f"- Поточна база: {BOARD_METADATA['current_base']}",
        f"- Старий Point A з XLSX: {BOARD_METADATA['old_point_a']}",
        f"- Старий Point B з XLSX: {BOARD_METADATA['old_point_b']}",
        f"- Ранній довгостроковий орієнтир з PDF: {BOARD_METADATA['older_pdf_goal']}",
        f"- Оновлений Point B: {BOARD_METADATA['updated_point_b']}",
        "",
        "## Короткий наратив для верхньої частини дошки",
        BOARD_METADATA["narrative"],
        "",
        "## Updated Point B",
        BOARD_METADATA["updated_point_b"],
        "",
        "## Проміжні цілі на 6-12 місяців",
    ]

    lines.extend(f"- {goal}" for goal in INTERMEDIATE_GOALS)
    lines.extend(["", "## Sticky notes", ""])

    for section in SECTIONS:
        lines.append(f"### {section['title']}")
        lines.extend(f"- {note}" for note in section["notes"])
        lines.append("")

    lines.extend(["## Файли для attachment cards на дошці", ""])
    for attachment in PRIMARY_ATTACHMENTS:
        lines.append(f"- {attachment['path']} - {attachment['purpose']}")

    lines.extend(["", "## Додаткові робочі джерела для формулювань", ""])
    lines.extend(f"- {reference}" for reference in SUPPORTING_REFERENCES)
    lines.extend(["", "## Короткі підписи біля attachment cards", ""])
    lines.extend(
        [
            "- PDF: раннє формулювання траєкторії, де важливі автономія, стабільність і ширший full-stack data напрям.",
            "- XLSX: точне джерело старої точки A/B, skill gaps і первинного ІПР; саме цей файл є головним evidence source для ретроспективи.",
            "",
            "## Submission checklist",
            "",
        ]
    )
    lines.extend(f"- {item}" for item in SUBMISSION_CHECKLIST)
    lines.extend(
        [
            "",
            "## Статус виконання в цьому середовищі",
            "- Локальний .jam-файл не є текстовим форматом для прямого редагування; він містить binary canvas payload.",
            "- Browser/Figma route зупинився на Figma login screen, тому live-оновлення дошки та реальний FigJam JPG export у цьому середовищі не виконані.",
            "- Натомість підготовлено paste-ready content package і локальний JPG mockup, щоб ручне завершення в FigJam було мінімальним.",
            "",
        ]
    )

    output_path.write_text("\n".join(lines), encoding="utf-8")


def render_board_image(output_path: Path) -> None:
    width, height = 3600, 2600
    image = Image.new("RGB", (width, height), "#F3EEE4")
    draw = ImageDraw.Draw(image)

    title_font = load_font(72, bold=True)
    subtitle_font = load_font(28)
    card_title_font = load_font(34, bold=True)
    card_body_font = load_font(24)
    section_title_font = load_font(40, bold=True)
    note_font = load_font(24)

    draw.rounded_rectangle((40, 34, width - 40, 270), radius=42, fill="#1F3742")
    draw.text((92, 84), BOARD_METADATA["title"], font=title_font, fill="#F7F4EE")
    draw.multiline_text(
        (96, 168),
        wrap_text(draw, BOARD_METADATA["subtitle"], subtitle_font, width - 220),
        font=subtitle_font,
        fill="#DCE8EB",
        spacing=8,
    )

    top_y0, top_y1 = 320, 820
    margin = 70
    gap = 28
    card_width = int((width - margin * 2 - gap * 2) / 3)

    trajectory_lines = [
        f"PDF (ранній горизонт): {BOARD_METADATA['older_pdf_goal']}",
        f"XLSX Point A: {BOARD_METADATA['old_point_a']}",
        f"XLSX Point B: {BOARD_METADATA['old_point_b']}",
        f"Поточна база: {BOARD_METADATA['current_base']}",
    ]
    draw_card(
        draw,
        (margin, top_y0, margin + card_width, top_y1),
        "Траєкторія",
        trajectory_lines,
        fill="#FFFDF8",
        outline="#D7CDBE",
        title_font=card_title_font,
        body_font=card_body_font,
    )

    goals_x0 = margin + card_width + gap
    goal_lines = [f"Updated Point B: {BOARD_METADATA['updated_point_b']}"] + [f"- {goal}" for goal in INTERMEDIATE_GOALS]
    draw_card(
        draw,
        (goals_x0, top_y0, goals_x0 + card_width, top_y1),
        "Нова точка B і проміжні цілі",
        goal_lines,
        fill="#FFFDF8",
        outline="#D7CDBE",
        title_font=card_title_font,
        body_font=card_body_font,
    )

    source_x0 = goals_x0 + card_width + gap
    source_lines = [
        "Обов'язкові attachment cards:",
        *[f"- {Path(item['path']).name}: {item['purpose']}" for item in PRIMARY_ATTACHMENTS],
        "",
        "Опора для формулювань:",
        *[f"- {Path(reference).name}" for reference in SUPPORTING_REFERENCES],
    ]
    draw_card(
        draw,
        (source_x0, top_y0, source_x0 + card_width, top_y1),
        "Джерела",
        source_lines,
        fill="#FFFDF8",
        outline="#D7CDBE",
        title_font=card_title_font,
        body_font=card_body_font,
    )

    columns_y0, columns_y1 = 880, height - 70
    column_gap = 22
    column_width = int((width - margin * 2 - column_gap * 4) / 5)

    for index, section in enumerate(SECTIONS):
        x0 = margin + index * (column_width + column_gap)
        x1 = x0 + column_width
        draw_sticky_section(
            draw,
            (x0, columns_y0, x1, columns_y1),
            section["title"],
            section["notes"],
            header_fill=section["header"],
            note_fill=section["fill"],
            title_font=section_title_font,
            body_font=note_font,
        )

    image.save(output_path, "JPEG", quality=92)


def main() -> None:
    ensure_output_dirs()

    pdf_path = WORKSPACE_DIR / "Формулювання кар'єрних цілей.pdf"
    workbook_path = WORKSPACE_DIR / "workshop - How to get a raise" / "Illya_Fefelov_SWOT_IPR_filled.xlsx"

    pdf_output = SOURCE_DIR / "career_goals_pdf_excerpt.txt"
    workbook_output = SOURCE_DIR / "swot_ipr_workbook_dump.txt"

    pdf_output.write_text(extract_pdf_text(pdf_path), encoding="utf-8")
    workbook_output.write_text(extract_workbook_text(workbook_path), encoding="utf-8")
    write_board_markdown(MARKDOWN_OUTPUT)
    render_board_image(IMAGE_OUTPUT)

    print(f"Wrote {pdf_output}")
    print(f"Wrote {workbook_output}")
    print(f"Wrote {MARKDOWN_OUTPUT}")
    print(f"Wrote {IMAGE_OUTPUT}")


if __name__ == "__main__":
    main()